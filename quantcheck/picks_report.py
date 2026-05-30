#!/usr/bin/env python3
"""
Fetch Quant GT monthly and weekly picks and export a designed Excel report.

Usage:
  python -m quantcheck.picks_report

Configuration is loaded from environment variables or .env.
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from quantcheck.scrape_parse import clean_text, extract_pick_date, rows_from_card_texts, rows_from_matrix

BASE = "https://quantgt.io"
ROOT = Path(os.environ.get("QUANTCHECK_HOME", Path(__file__).resolve().parents[1]))
OUT_DIR = ROOT / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)
STATE_DIR = ROOT
STATE_DIR.mkdir(parents=True, exist_ok=True)

PROFILE = ROOT / "browser-profile"
PROFILE.mkdir(parents=True, exist_ok=True)

EMAIL = os.environ.get("QUANTGT_EMAIL", "")
PASSWORD = os.environ.get("QUANTGT_PASSWORD", "")

# Quant GT light theme palette
BG = "FFFFFF"          # default white worksheet background
SURFACE = "FFFFFF"     # table body white
SURFACE_2 = "FFFFFF"   # no alternate color; user wants all white
GREEN = "16A34A"       # Quant GT green
GREEN_DARK = "0F7A36"
GREEN_SOFT = "DCFCE7"
TEXT = "0F172A"        # slate-900
MUTED = "64748B"       # slate-500
GRID = "D7E3DA"
RED = "DC2626"
AMBER = "D97706"


def is_login_prompt_visible(page) -> bool:
    """Return true only for real auth prompts, not marketing copy mentioning sign in."""
    try:
        if page.locator('input[type="email"], input[type="password"]').count() > 0:
            return True
        return page.get_by_role("button", name=re.compile(r"^(log in|sign in)$", re.I)).count() > 0
    except Exception:
        return False


def has_auth_session(page) -> bool:
    try:
        cookies = page.context.cookies(BASE)
        return any(c.get('name') == '__Secure-authjs.session-token' for c in cookies)
    except Exception:
        return False


def has_picks_content(page) -> bool:
    try:
        return bool(page.evaluate(
            """() => {
              const text = (document.querySelector('main')?.innerText || document.body.innerText || '').replace(/\\s+/g, ' ');
              const tableRows = document.querySelectorAll('table tbody tr').length;
              const ariaRows = document.querySelectorAll('[role="row"] [role="cell"], [role="row"] [role="gridcell"]').length;
              return tableRows > 0 || ariaRows > 0 || /\\bGT\\s*Score\\b/i.test(text);
            }"""
        ))
    except Exception:
        return False


def wait_for_picks_content(page, timeout: int = 20000) -> None:
    page.wait_for_function(
        """() => {
          const text = (document.querySelector('main')?.innerText || document.body.innerText || '').replace(/\\s+/g, ' ');
          const tableRows = document.querySelectorAll('table tbody tr').length;
          const ariaRows = document.querySelectorAll('[role="row"] [role="cell"], [role="row"] [role="gridcell"]').length;
          return tableRows > 0 || ariaRows > 0 || /\\bGT\\s*Score\\b/i.test(text);
        }""",
        timeout=timeout,
    )


def wait_for_parsable_picks_rows(page, mode: str, attempts: int = 3, timeout: int = 20000) -> List[Dict[str, Any]]:
    """Wait until the rendered page can be parsed into real pick rows.

    Quant GT sometimes renders headings/labels like "GT Score" before the actual
    table/card rows finish hydrating. Treating that as success caused occasional
    zero-row Monthly captures. This helper requires parsed rows, and reloads the
    page before retrying so scheduled runs self-heal instead of immediately
    falling into the validation gate.
    """
    errors = []
    attempts = max(1, int(attempts or 1))
    for attempt in range(1, attempts + 1):
        try:
            wait_for_picks_content(page, timeout=timeout)
            # Give client-side table/card hydration one short extra beat after
            # the first content signal, then parse real rows.
            page.wait_for_timeout(1200)
            rows = rows_from_table(page, mode)
            if rows:
                return rows
            text = clean_text(page.locator("main").inner_text(timeout=3000))[:500]
            errors.append(f"attempt {attempt}: no parsed {mode} rows; main={text!r}")
        except Exception as e:
            errors.append(f"attempt {attempt}: {type(e).__name__}: {e}")
        if attempt < attempts:
            try:
                page.reload(wait_until="domcontentloaded", timeout=45000)
                try:
                    page.wait_for_load_state("load", timeout=15000)
                except PlaywrightTimeoutError:
                    pass
                page.wait_for_timeout(2500)
            except Exception as e:
                errors.append(f"attempt {attempt}: reload failed: {type(e).__name__}: {e}")
    raise RuntimeError(f"{mode} page did not produce parsable picks rows after {attempts} attempts: " + " | ".join(errors[-3:]))


def login(page):
    # `networkidle` is brittle on Quant GT because embedded market/news widgets
    # can keep polling. Use document readiness + table hydration instead.
    page.goto(f"{BASE}/dashboard/quantgt-picks", wait_until="domcontentloaded", timeout=45000)
    try:
        page.wait_for_load_state("load", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(2500)
    if has_picks_content(page) and has_auth_session(page) and not is_login_prompt_visible(page):
        return
    page.goto(f"{BASE}/login?redirect=/dashboard/quantgt-picks", wait_until="domcontentloaded", timeout=45000)
    try:
        page.wait_for_load_state("load", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(1000)
    if page.get_by_role("button", name=re.compile("log in|sign in", re.I)).count() > 0:
        page.get_by_role("button", name=re.compile("log in|sign in", re.I)).first.click()
    email_box = page.get_by_placeholder("you@example.com") if page.get_by_placeholder("you@example.com").count() else page.locator('input[type="email"]')
    pass_box = page.get_by_placeholder("min. 8 characters") if page.get_by_placeholder("min. 8 characters").count() else page.locator('input[type="password"]')
    email_box.fill(EMAIL)
    pass_box.fill(PASSWORD)
    page.get_by_role("button", name=re.compile("log in|sign in", re.I)).first.click()
    try:
        page.wait_for_load_state("domcontentloaded", timeout=20000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(3000)
    page.goto(f"{BASE}/dashboard/quantgt-picks", wait_until="domcontentloaded", timeout=45000)
    try:
        page.wait_for_load_state("load", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(3000)
    wait_for_picks_content(page)
    if not has_picks_content(page) or is_login_prompt_visible(page):
        raise RuntimeError("Login did not complete or picks table not visible")


def assert_authenticated_page(page, label: str):
    """Reject unauthenticated/demo pages before their content can enter monitor state."""
    if is_login_prompt_visible(page):
        raise RuntimeError(f"{label} page is not authenticated: login prompt visible")
    cookies = page.context.cookies(BASE)
    names = {c.get('name') for c in cookies}
    if '__Secure-authjs.session-token' not in names:
        raise RuntimeError(f"{label} page is not authenticated: missing auth session cookie")
    if not has_picks_content(page):
        raise RuntimeError(f"{label} page has no picks content")


def rows_from_table(page, mode: str) -> List[Dict[str, Any]]:
    js = r"""
    (mode) => {
      const clean = s => (s || '').trim().replace(/\s+/g, ' ');
      const tableRows = [...document.querySelectorAll('table tr')]
        .map(tr => [...tr.querySelectorAll('th,td')].map(td => clean(td.innerText || td.textContent)))
        .filter(row => row.some(Boolean));
      const ariaRows = [...document.querySelectorAll('[role="row"]')]
        .map(tr => [...tr.querySelectorAll('[role="columnheader"],[role="cell"],[role="gridcell"]')].map(td => clean(td.innerText || td.textContent)))
        .filter(row => row.some(Boolean));
      const cards = [...document.querySelectorAll('main article, main [data-slot*="card"], main [class*="card"], main [class*="Card"]')]
        .map(el => clean(el.innerText || el.textContent))
        .filter(text => text && /GT\s*Score|Rating|Sector|Held Since|Return/i.test(text));
      return {matrix: tableRows.length ? tableRows : ariaRows, cards};
    }
    """
    payload = page.evaluate(js, mode)
    rows = rows_from_matrix(payload.get("matrix") or [], mode)
    if rows:
        return rows
    return rows_from_card_texts(payload.get("cards") or [], mode)


def extract_details_for_visible_expanded(page) -> Dict[str, str]:
    # Expanded detail row is the row whose first td starts with a price like $169.19.
    js = r"""
    () => {
      const rows = [...document.querySelectorAll('table tbody tr')];
      for (const tr of rows) {
        const cells = [...tr.querySelectorAll('td')].map(td => td.innerText.trim());
        if (cells.length && cells[0].startsWith('$')) {
          const txt = cells[0].replace(/\s+/g, ' ');
          const get = (label, nextLabels) => {
            const i = txt.indexOf(label);
            if (i < 0) return '';
            let start = i + label.length;
            let end = txt.length;
            for (const nl of nextLabels) {
              const j = txt.indexOf(nl, start);
              if (j >= 0 && j < end) end = j;
            }
            return txt.slice(start, end).trim();
          };
          return {
            raw: txt,
            current_price: (txt.match(/^\$[0-9.,]+/)||[''])[0],
            chart_return: (txt.match(/([+-][0-9.]+%) ·/)||['',''])[1],
            buy_or_entry_price: get('Buy price:', ['P/E (TTM)', 'Market Cap']) || get('Entry price:', ['P/E (TTM)', 'Market Cap']),
            pe_ttm: get('P/E (TTM)', ['Market Cap']),
            market_cap: get('Market Cap', ['Revenue (TTM)']),
            revenue_ttm: get('Revenue (TTM)', ['Revenue Growth (YoY)']),
            revenue_growth_yoy: get('Revenue Growth (YoY)', ['Next Earnings']),
            next_earnings: get('Next Earnings', ['Analyst Signal']),
            analyst_signal: get('Analyst Signal', ['Momentum']),
            momentum: get('Momentum', ['Relative Strength']),
            relative_strength: get('Relative Strength', ['Sandisk', 'Applied', 'Lumentum', 'Viavi', 'Ciena', 'FORM', 'DigitalOcean', 'Planet', 'Western', 'Corning'])
          };
        }
      }
      return {};
    }
    """
    return page.evaluate(js)


def expand_and_attach_details(page, rows: List[Dict[str, Any]], mode: str) -> List[Dict[str, Any]]:
    # Expand each data row by symbol and extract metrics from the detail row immediately below it.
    detail_js = r"""
    async (sym) => {
      const sleep = ms => new Promise(r => setTimeout(r, ms));
      const findRow = () => [...document.querySelectorAll('table tbody tr')]
        .find(tr => [...tr.querySelectorAll('td')].some(td => td.innerText.trim() === sym));
      let row = findRow();
      if (!row) return {detail_error: 'row not found'};
      let detail = row.nextElementSibling;
      if (!detail || !detail.innerText.trim().startsWith('$')) {
        row.click();
        await sleep(900);
        row = findRow();
        detail = row ? row.nextElementSibling : null;
      }
      if (!detail || !detail.innerText.trim().startsWith('$')) return {detail_error: 'detail row not found'};
      const txt = detail.innerText.trim().replace(/\s+/g, ' ');
      const get = (label, nextLabels) => {
        const i = txt.indexOf(label);
        if (i < 0) return '';
        const start = i + label.length;
        let end = txt.length;
        for (const nl of nextLabels) {
          const j = txt.indexOf(nl, start);
          if (j >= 0 && j < end) end = j;
        }
        return txt.slice(start, end).trim();
      };
      const out = {
        current_price: (txt.match(/^\$[0-9.,]+/)||[''])[0],
        chart_return: (txt.match(/([+-][0-9.]+%) ·/)||['',''])[1],
        buy_or_entry_price: get('Buy price:', ['P/E (TTM)', 'Market Cap']) || get('Entry price:', ['P/E (TTM)', 'Market Cap']) || ((txt.match(new RegExp('\\b' + sym.replace(/[.*+?^${}()|[\\]\\]/g, '\\$&') + '\\s*:\\s*(\\$[0-9.,]+)')) || ['', ''])[1]),
        pe_ttm: get('P/E (TTM)', ['Market Cap']),
        market_cap: get('Market Cap', ['Revenue (TTM)']),
        revenue_ttm: get('Revenue (TTM)', ['Revenue Growth (YoY)']),
        revenue_growth_yoy: get('Revenue Growth (YoY)', ['Next Earnings']),
        next_earnings: get('Next Earnings', ['Analyst Signal']),
        analyst_signal: get('Analyst Signal', ['Momentum']),
        momentum: get('Momentum', ['Relative Strength']),
        relative_strength: get('Relative Strength', [])
      };
      // Collapse after extraction so the next lookup is clean.
      row = findRow();
      if (row && row.nextElementSibling && row.nextElementSibling.innerText.trim().startsWith('$')) row.click();
      await sleep(150);
      return out;
    }
    """
    for r in rows:
        sym = r.get("symbol")
        if not sym:
            continue
        try:
            details = page.evaluate(detail_js, sym)
            if details:
                r.update({k: clean_text(v) for k, v in details.items()})
        except Exception as e:
            r["detail_error"] = f"detail not captured: {type(e).__name__}"
    return rows


def fetch():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 1440, "height": 1200}, locale="en-US", storage_state=None)
        page = context.new_page()
        login(page)

        page.goto(f"{BASE}/dashboard/quantgt-picks", wait_until="domcontentloaded", timeout=45000)
        try:
            page.wait_for_load_state("load", timeout=15000)
        except PlaywrightTimeoutError:
            pass
        page.wait_for_timeout(2500)
        assert_authenticated_page(page, "monthly")
        # Sometimes the authenticated picks content hydrates after document load.
        wait_for_picks_content(page)
        monthly_rows = wait_for_parsable_picks_rows(page, "monthly")
        monthly_date_text = clean_text(page.locator("main").inner_text())
        monthly_pick_date = extract_pick_date(monthly_date_text, "monthly")
        monthly_rows = expand_and_attach_details(page, monthly_rows, "monthly")

        page.goto(f"{BASE}/dashboard/weekly-picks", wait_until="domcontentloaded", timeout=45000)
        try:
            page.wait_for_load_state("load", timeout=15000)
        except PlaywrightTimeoutError:
            pass
        page.wait_for_timeout(2500)
        assert_authenticated_page(page, "weekly")
        wait_for_picks_content(page)
        weekly_rows = wait_for_parsable_picks_rows(page, "weekly")
        main_text = clean_text(page.locator("main").inner_text())
        weekly_pick_date = extract_pick_date(main_text, "weekly")
        weekly_rows = expand_and_attach_details(page, weekly_rows, "weekly")

        browser.close()

    return {
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "source": BASE,
        "monthly": {"page": f"{BASE}/dashboard/quantgt-picks", "pick_date": monthly_pick_date, "rows": monthly_rows},
        "weekly": {"page": f"{BASE}/dashboard/weekly-picks", "pick_date": weekly_pick_date, "rows": weekly_rows},
    }


def gt_score_float(v: str):
    m = re.search(r"[0-9.]+", v or "")
    return float(m.group(0)) if m else None


def pct_float(v: str):
    m = re.search(r"[-+]?[0-9.]+", v or "")
    return float(m.group(0)) if m else None


def setup_sheet(ws, title, subtitle):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    # Keep worksheet default white; only style headers and data cards.
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=TEXT)
    ws["A1"].fill = PatternFill(fill_type=None)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=11, color=MUTED)
    ws["A2"].fill = PatternFill(fill_type=None)
    ws.row_dimensions[2].height = 24


def style_table(ws, headers, start_row, widths):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(start_row, col, h)
        cell.fill = PatternFill("solid", fgColor=GREEN_SOFT)
        cell.font = Font(name="Aptos", size=10, bold=True, color=GREEN_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=GREEN), top=Side(style="thin", color=GRID))
    ws.row_dimensions[start_row].height = 24
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_body_style(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        fill = SURFACE if r % 2 else SURFACE_2
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color=GRID))
    for row in range(min_row, max_row + 1):
        ws.row_dimensions[row].height = 28


def set_alignments(ws, align_map, min_row, max_row):
    """align_map: {column_number: horizontal_alignment}"""
    for c, align in align_map.items():
        for r in range(min_row, max_row + 1):
            ws.cell(r, c).alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def set_print_layout(ws, last_col, last_row):
    ws.freeze_panes = "A7"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:6"
    ws.auto_filter.ref = f"A6:{get_column_letter(last_col)}{last_row}"
    ws.sheet_view.zoomScale = 95


def write_summary(wb, data):
    ws = wb.active
    ws.title = "Overview"
    setup_sheet(ws, "Quant GT Picks Report", f"Fetched at {data['fetched_at']} · Source: quantgt.io")

    # Overview should only show the actual monthly/weekly pick lists, with update dates beside titles.
    ws.merge_cells("A6:C6")
    ws.merge_cells("E6:G6")
    ws.cell(6, 1, f"Monthly Picks · {data['monthly']['pick_date']}")
    ws.cell(6, 5, f"Weekly Picks · {data['weekly']['pick_date']}")
    for c in [1, 5]:
        ws.cell(6, c).fill = PatternFill("solid", fgColor=GREEN_SOFT)
        ws.cell(6, c).font = Font(name="Aptos", bold=True, color=GREEN_DARK)
        ws.cell(6, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6, c).border = Border(bottom=Side(style="medium", color=GREEN))

    ws.cell(7, 1, "Symbol")
    ws.cell(7, 2, "GT Score")
    ws.cell(7, 3, "Return")
    ws.cell(7, 5, "Symbol")
    ws.cell(7, 6, "GT Score")
    ws.cell(7, 7, "Signal")
    for c in [1, 2, 3, 5, 6, 7]:
        ws.cell(7, c).fill = PatternFill("solid", fgColor=GREEN_SOFT)
        ws.cell(7, c).font = Font(name="Aptos", bold=True, color=GREEN_DARK)
        ws.cell(7, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(7, c).border = Border(bottom=Side(style="thin", color=GRID))

    for i, r in enumerate(data["monthly"]["rows"], 8):
        ws.cell(i, 1, r.get("symbol"))
        ws.cell(i, 2, r.get("gt_score"))
        ws.cell(i, 3, r.get("return", ""))
    for i, r in enumerate(data["weekly"]["rows"], 8):
        ws.cell(i, 5, r.get("symbol"))
        ws.cell(i, 6, r.get("gt_score"))
        ws.cell(i, 7, r.get("analyst_signal"))

    max_row = max(8 + len(data['weekly']['rows']) - 1, 8 + len(data['monthly']['rows']) - 1)
    apply_body_style(ws, 8, max_row, 7)
    for rr in range(8, max_row + 1):
        for cc in [1, 5]:
            ws.cell(rr, cc).font = Font(name="Aptos", size=10, bold=True, color=GREEN)
        val = str(ws.cell(rr, 3).value or '')
        if val.startswith('+'):
            ws.cell(rr, 3).font = Font(name="Aptos", size=10, bold=True, color=GREEN)
        elif val.startswith('-'):
            ws.cell(rr, 3).font = Font(name="Aptos", size=10, bold=True, color=RED)

    # Logical alignment: identifiers left, scores/returns centered, labels left.
    set_alignments(ws, {1: "left", 2: "center", 3: "center", 5: "left", 6: "center", 7: "left"}, 7, max_row)
    widths = {1: 12, 2: 12, 3: 12, 4: 4, 5: 12, 6: 12, 7: 16}
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w
    set_print_layout(ws, 7, max_row)


def write_picks_sheet(wb, sheet_name, title, pick_date, rows, mode):
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, title, f"Recommendation date: {pick_date}")
    if mode == "monthly":
        headers = ["Rank", "Symbol", "Company", "Held Since", "Return", "Sector", "GT Score", "Current Price", "Entry/Buy Price", "Revenue Growth", "Next Earnings", "Analyst Signal"]
        widths = [8, 10, 30, 13, 12, 24, 12, 14, 16, 16, 18, 18]
    else:
        headers = ["Rank", "Symbol", "Company", "Sector", "GT Score", "Current Price", "Buy Price", "Market Cap", "Revenue Growth", "Next Earnings", "Analyst Signal"]
        widths = [8, 10, 30, 24, 12, 14, 16, 16, 16, 18, 18]
    style_table(ws, headers, 6, widths)
    for idx, r in enumerate(rows, 1):
        rr = 6 + idx
        if mode == "monthly":
            vals = [idx, r.get("symbol"), r.get("company"), r.get("held_since"), r.get("return"), r.get("sector"), r.get("gt_score"), r.get("current_price"), r.get("buy_or_entry_price"), r.get("revenue_growth_yoy"), r.get("next_earnings"), r.get("analyst_signal")]
        else:
            vals = [idx, r.get("symbol"), r.get("company"), r.get("sector"), r.get("gt_score"), r.get("current_price"), r.get("buy_or_entry_price"), r.get("market_cap"), r.get("revenue_growth_yoy"), r.get("next_earnings"), r.get("analyst_signal")]
        for c, v in enumerate(vals, 1):
            ws.cell(rr, c, v)
    if rows:
        apply_body_style(ws, 7, 6 + len(rows), len(headers))
        data_last_row = 6 + len(rows)
    else:
        data_last_row = 6
    # Logical alignment by field type.
    if mode == "monthly":
        align_map = {
            1: "center", 2: "left", 3: "left", 4: "center", 5: "center", 6: "left",
            7: "center", 8: "right", 9: "right", 10: "center", 11: "center", 12: "center"
        }
    else:
        align_map = {
            1: "center", 2: "left", 3: "left", 4: "left", 5: "center", 6: "right",
            7: "right", 8: "right", 9: "center", 10: "center", 11: "center"
        }
    set_alignments(ws, align_map, 7, data_last_row)
    # Accent key columns
    for rr in range(7, 7 + len(rows)):
        for cc in [2, 5 if mode == 'weekly' else 7]:
            ws.cell(rr, cc).font = Font(name="Aptos", size=10, bold=True, color=GREEN)
        # Return/revenue growth coloring
        for cc in range(1, len(headers)+1):
            val = str(ws.cell(rr, cc).value or '')
            if val.startswith('+'):
                ws.cell(rr, cc).font = Font(name="Aptos", size=10, bold=True, color=GREEN)
            elif val.startswith('-'):
                ws.cell(rr, cc).font = Font(name="Aptos", size=10, bold=True, color=RED)
    set_print_layout(ws, len(headers), data_last_row)


def export_excel(data) -> Path:
    wb = Workbook()
    write_summary(wb, data)
    write_picks_sheet(wb, "Monthly Picks", "Monthly Picks", data["monthly"]["pick_date"], data["monthly"]["rows"], "monthly")
    write_picks_sheet(wb, "Weekly Picks", "Weekly Picks", data["weekly"]["pick_date"], data["weekly"]["rows"], "weekly")

    path = OUT_DIR / f"quantgt_picks_report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path


def main():
    global EMAIL, PASSWORD
    EMAIL = os.environ.get('QUANTGT_EMAIL', EMAIL)
    PASSWORD = os.environ.get('QUANTGT_PASSWORD', PASSWORD)
    data = fetch()
    state_path = STATE_DIR / "latest_picks.json"
    state_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx = export_excel(data)
    print(json.dumps({
        "excel": str(xlsx),
        "state": str(state_path),
        "monthly_date": data["monthly"]["pick_date"],
        "monthly_count": len(data["monthly"]["rows"]),
        "weekly_date": data["weekly"]["pick_date"],
        "weekly_count": len(data["weekly"]["rows"]),
        "monthly_symbols": [r.get("symbol") for r in data["monthly"]["rows"]],
        "weekly_symbols": [r.get("symbol") for r in data["weekly"]["rows"]],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
